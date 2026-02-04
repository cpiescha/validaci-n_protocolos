# -*- coding: utf-8 -*-
import argparse
import re
import unicodedata
from dataclasses import dataclass
from typing import Any, List, Optional, Tuple

import openpyxl as ox
import pdfplumber
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    return text


def extract_numbers(text: str) -> List[float]:
    if not text:
        return []
    cleaned = text.replace(",", ".")
    nums = re.findall(r"[-+]?\d+(?:\.\d+)?", cleaned)
    out = []
    for n in nums:
        try:
            out.append(float(n))
        except ValueError:
            pass
    return out


def find_excel_value(ws, label_substring: str, col_value: int = 2) -> Optional[Any]:
    needle = norm_text(label_substring)
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=10):
        cell = row[0]
        if isinstance(cell.value, str):
            if needle in norm_text(cell.value):
                return ws.cell(cell.row, col_value).value
    return None


def find_excel_value_in_col_a(ws, regex_pattern: str, col_value: int = 2) -> Optional[Any]:
    pattern = re.compile(regex_pattern, re.IGNORECASE)
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=10):
        cell = row[0]
        if isinstance(cell.value, str) and pattern.search(cell.value):
            return ws.cell(cell.row, col_value).value
    return None


def excel_tipo_to_pdf(tipo: Any) -> str:
    t = norm_text(tipo).replace(" ", "")
    if "CRHV3" in t:
        return "SECO"
    if "OTHV3" in t:
        return "ACEITE"
    return t


def parse_pdf_text(pdf_path: str) -> str:
    text_parts: List[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text_parts.append(page.extract_text() or "")
            for settings in [
                {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
                {"vertical_strategy": "lines", "horizontal_strategy": "text"},
                {"vertical_strategy": "text", "horizontal_strategy": "lines"},
                {"vertical_strategy": "text", "horizontal_strategy": "text"},
            ]:
                try:
                    tables = page.extract_tables(settings)
                except Exception:
                    tables = []
                for table in tables or []:
                    for row in table:
                        if row:
                            text_parts.append(" ".join(cell for cell in row if cell))
    return "\n".join(text_parts)


def find_pdf_group(text_norm: str) -> Optional[str]:
    m = re.search(r"GRUPO DE CONEXION\s*:?\s*([A-Z0-9]+)", text_norm)
    if m:
        return m.group(1)
    m = re.search(r"FASES\s*:\s*\d+\s*([A-Z0-9]+)", text_norm)
    if m:
        return m.group(1)
    return None


def find_pdf_kv_pair(text: str) -> Optional[Tuple[float, float]]:
    pairs = re.findall(r"(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)", text)
    candidates = []
    for a, b in pairs:
        try:
            a_f, b_f = float(a), float(b)
            if 1.0 <= a_f <= 50 and 0.05 <= b_f <= 5:
                candidates.append((a_f, b_f))
        except ValueError:
            continue
    return candidates[0] if candidates else None


def find_value_in_lines(text: str, key_norm: str) -> Optional[float]:
    for line in text.splitlines():
        if key_norm in norm_text(line):
            nums = extract_numbers(line)
            if nums:
                return nums[-1]
    return None


def find_code_producto(text_norm: str) -> Optional[str]:
    m = re.search(r"CODIGO DEL PRODUCTO\s*:?\s*([0-9]+)", text_norm)
    return m.group(1) if m else None


def find_impedance_value(text_norm: str) -> Optional[float]:
    idx = text_norm.find("IMPEDANCIA @ 120")
    if idx == -1:
        return None
    segment = text_norm[idx : idx + 400]
    end = segment.find("NIVEL DE RUIDO")
    if end != -1:
        segment = segment[:end]
    nums = extract_numbers(segment)
    if nums and nums[0] == 120:
        nums = nums[1:]
    if nums:
        return nums[-1]
    return None


def normalize_id(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    m = re.fullmatch(r"(\d+)(?:\.0+)?", s)
    if m:
        return m.group(1)
    return s


def to_kv(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        v = float(str(value).replace(",", "."))
    except ValueError:
        return None
    if v >= 100:
        return v / 1000.0
    return v


def normalize_group(value: Any) -> str:
    return norm_text(value).replace(" ", "")


@dataclass
class CompareResult:
    field: str
    excel_value: Any
    pdf_value: Any
    ok: bool
    details: str = ""


def compare_text(field: str, excel_value: Any, pdf_value: Any) -> CompareResult:
    ev = norm_text(excel_value)
    pv = norm_text(pdf_value)
    ok = ev == pv and ev != ""
    return CompareResult(field, excel_value, pdf_value, ok)


def compare_ids(field: str, excel_value: Any, pdf_value: Any) -> CompareResult:
    ev = normalize_id(excel_value)
    pv = normalize_id(pdf_value)
    ok = ev == pv and ev != ""
    return CompareResult(field, ev, pv, ok)


def compare_group(field: str, excel_value: Any, pdf_value: Any) -> CompareResult:
    ev = normalize_group(excel_value)
    pv = normalize_group(pdf_value)
    ok = ev == pv and ev != ""
    return CompareResult(field, excel_value, pdf_value, ok)


def compare_numeric(field: str, excel_value: Any, pdf_value: Any, unit: str, rel_tol: float, abs_tol: float) -> CompareResult:
    if excel_value is None or pdf_value is None:
        return CompareResult(field, excel_value, pdf_value, False, "missing")
    try:
        ev = float(str(excel_value).replace(",", "."))
        pv = float(str(pdf_value).replace(",", "."))
    except ValueError:
        return CompareResult(field, excel_value, pdf_value, False, "invalid number")

    diff = abs(ev - pv)
    ok = diff <= max(abs_tol, rel_tol * abs(ev))
    details = f"diff={diff:.6g} {unit}"
    return CompareResult(field, ev, pv, ok, details)


def run_validation(excel_path: str, pdf_path: str, sheet: str) -> List[CompareResult]:
    wb = ox.load_workbook(excel_path, data_only=True)
    ws = wb[sheet]

    excel = {
        "cp": find_excel_value(ws, "No.  C.P"),
        "kva": find_excel_value(ws, "KVA"),
        "tipo": find_excel_value(ws, "TIPO"),
        "v_prim": find_excel_value(ws, "V. PRIM."),
        "v_sec": find_excel_value(ws, "V. SEC."),
        "grupo": find_excel_value(ws, "GRUPO DE CONEXION"),
        "po": find_excel_value_in_col_a(ws, r"PO\s+GARANTIZADO"),
        "io": find_excel_value_in_col_a(ws, r"IO\s+GARANTIZADO"),
        "pcc": find_excel_value_in_col_a(ws, r"PCC\s+GARANTIZADO"),
        "uz": find_excel_value_in_col_a(ws, r"UZ\s+GARANTIZADO"),
    }

    pdf_text = parse_pdf_text(pdf_path)
    pdf_text_norm = norm_text(pdf_text)

    pdf = {
        "codigo_producto": None,
        "tipo": None,
        "grupo": None,
        "kva": None,
        "v_prim_kv": None,
        "v_sec_kv": None,
        "po_kw": None,
        "pcc_kw": None,
        "io_pct": None,
        "uz_pct": None,
    }

    pdf["codigo_producto"] = find_code_producto(pdf_text_norm)

    m = re.search(r"TIPO\s*:?\s*([A-Z]+)", pdf_text_norm)
    pdf["tipo"] = m.group(1) if m else None

    pdf["grupo"] = find_pdf_group(pdf_text_norm)

    kv_pair = find_pdf_kv_pair(pdf_text)
    if kv_pair:
        pdf["v_prim_kv"], pdf["v_sec_kv"] = kv_pair

    m = re.search(r"POTENCIA \[KVA\].*?(\d+(?:[\.,]\d+)?)", pdf_text_norm)
    if m:
        pdf["kva"] = float(m.group(1).replace(",", "."))

    m = re.search(r"PERDIDAS A VACIO \[KW\]\s*([0-9]+(?:[\.,][0-9]+)?)", pdf_text_norm)
    if m:
        pdf["po_kw"] = float(m.group(1).replace(",", "."))

    m = re.search(r"PERDIDAS EN CORTOCIRCUITO \[KW\]\s*([0-9]+(?:[\.,][0-9]+)?)", pdf_text_norm)
    if m:
        pdf["pcc_kw"] = float(m.group(1).replace(",", "."))

    m = re.search(r"CORRIENTE DE EXCITACION \[%\]\s*([0-9]+(?:[\.,][0-9]+)?)", pdf_text_norm)
    if m:
        pdf["io_pct"] = float(m.group(1).replace(",", "."))

    pdf["uz_pct"] = find_impedance_value(pdf_text_norm)

    results: List[CompareResult] = []

    results.append(compare_ids("CP vs Codigo del producto", excel["cp"], pdf["codigo_producto"]))
    results.append(compare_text("Tipo", excel_tipo_to_pdf(excel["tipo"]), pdf["tipo"]))
    results.append(compare_group("Grupo de conexion", excel["grupo"], pdf["grupo"]))

    results.append(compare_numeric(
        "V. Primario [kV]",
        to_kv(excel["v_prim"]),
        pdf["v_prim_kv"],
        "kV",
        rel_tol=0.01,
        abs_tol=0.01,
    ))
    results.append(compare_numeric(
        "V. Secundario [kV]",
        to_kv(excel["v_sec"]),
        pdf["v_sec_kv"],
        "kV",
        rel_tol=0.01,
        abs_tol=0.01,
    ))

    results.append(compare_numeric(
        "Potencia [kVA]",
        excel["kva"],
        pdf["kva"],
        "kVA",
        rel_tol=0.01,
        abs_tol=1.0,
    ))

    results.append(compare_numeric(
        "Po garantizado [W]",
        excel["po"],
        (pdf["po_kw"] * 1000.0) if pdf["po_kw"] is not None else None,
        "W",
        rel_tol=0.02,
        abs_tol=50.0,
    ))
    results.append(compare_numeric(
        "Pcc garantizado [W]",
        excel["pcc"],
        (pdf["pcc_kw"] * 1000.0) if pdf["pcc_kw"] is not None else None,
        "W",
        rel_tol=0.02,
        abs_tol=50.0,
    ))

    results.append(compare_numeric(
        "Io garantizado [%]",
        excel["io"],
        pdf["io_pct"],
        "%",
        rel_tol=0.05,
        abs_tol=0.1,
    ))
    results.append(compare_numeric(
        "Uz garantizado [%]",
        excel["uz"],
        pdf["uz_pct"],
        "%",
        rel_tol=0.05,
        abs_tol=0.1,
    ))

    return results


def format_results(results: List[CompareResult]) -> str:
    lines = []
    lines.append("VALIDACION EXCEL vs PDF")
    lines.append("-" * 60)
    for r in results:
        status = "OK" if r.ok else "NO"
        lines.append(f"{status:>2} | {r.field}: Excel={r.excel_value} | PDF={r.pdf_value} {r.details}")
    ok_count = sum(1 for r in results if r.ok)
    lines.append("-" * 60)
    lines.append(f"Resumen: {ok_count}/{len(results)} OK")
    return "\n".join(lines)


def main_cli() -> int:
    parser = argparse.ArgumentParser(description="Valida datos entre Excel y PDF")
    parser.add_argument("--excel", default="protocolo.xlsm")
    parser.add_argument("--pdf", default="datos.pdf")
    parser.add_argument("--sheet", default="DATOS")
    args = parser.parse_args()

    results = run_validation(args.excel, args.pdf, args.sheet)
    print(format_results(results))
    return 0


def launch_gui() -> None:
    root = tk.Tk()
    root.title("Validador Excel vs PDF")
    root.geometry("840x520")

    excel_var = tk.StringVar(value="protocolo.xlsm")
    pdf_var = tk.StringVar(value="datos.pdf")
    sheet_var = tk.StringVar(value="DATOS")

    def pick_excel() -> None:
        path = filedialog.askopenfilename(
            title="Seleccionar Excel",
            filetypes=[("Excel", "*.xlsm *.xlsx *.xls")],
        )
        if path:
            excel_var.set(path)

    def pick_pdf() -> None:
        path = filedialog.askopenfilename(
            title="Seleccionar PDF",
            filetypes=[("PDF", "*.pdf")],
        )
        if path:
            pdf_var.set(path)

    def run_and_log() -> None:
        excel_path = excel_var.get().strip()
        pdf_path = pdf_var.get().strip()
        sheet = sheet_var.get().strip() or "DATOS"
        if not excel_path or not pdf_path:
            messagebox.showwarning("Faltan archivos", "Selecciona el Excel y el PDF.")
            return
        try:
            results = run_validation(excel_path, pdf_path, sheet)
            log.delete("1.0", tk.END)
            log.insert(tk.END, format_results(results))
        except Exception as exc:
            log.delete("1.0", tk.END)
            log.insert(tk.END, f"Error: {exc}")

    frm = tk.Frame(root)
    frm.pack(fill=tk.BOTH, expand=False, padx=12, pady=12)

    tk.Label(frm, text="Archivo Excel:").grid(row=0, column=0, sticky="w")
    tk.Entry(frm, textvariable=excel_var, width=80).grid(row=0, column=1, padx=6, pady=4)
    tk.Button(frm, text="Buscar...", command=pick_excel).grid(row=0, column=2, padx=6)

    tk.Label(frm, text="Archivo PDF:").grid(row=1, column=0, sticky="w")
    tk.Entry(frm, textvariable=pdf_var, width=80).grid(row=1, column=1, padx=6, pady=4)
    tk.Button(frm, text="Buscar...", command=pick_pdf).grid(row=1, column=2, padx=6)

    tk.Label(frm, text="Hoja Excel:").grid(row=2, column=0, sticky="w")
    tk.Entry(frm, textvariable=sheet_var, width=20).grid(row=2, column=1, sticky="w", padx=6, pady=4)
    tk.Button(frm, text="Validar", command=run_and_log).grid(row=2, column=2, padx=6)

    log = ScrolledText(root, wrap=tk.WORD, height=20)
    log.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))

    root.mainloop()


if __name__ == "__main__":
    # If user provides CLI args, run in CLI mode. Otherwise launch GUI.
    import sys

    if len(sys.argv) > 1:
        raise SystemExit(main_cli())
    launch_gui()
